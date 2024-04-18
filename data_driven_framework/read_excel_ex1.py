import openpyxl

wb=openpyxl.load_workbook("C:\\Users\\care\\Desktop\\ddf.xlsx")
sh=wb['Sheet1']
totalrows=sh.max_row
print(totalrows)
totalcolumn=sh.max_column
print(totalcolumn)
data=sh.cell(row=4,column=2).value
print(data)
data1=sh.cell(row=6,column=1).value
print(data1)

print("=======================================================")

for r in range(1,totalrows+1):
    for c in range(1,totalcolumn+1):
        data2=sh.cell(row=r,column=c).value
        print(data2)
