import openpyxl
path="C:\\Users\\care\\Desktop\\ddf.xlsx"
wb=openpyxl.load_workbook(path)
sh=wb['Sheet2']
#sh.cell(row=1,column=1).value='python'
for r in range(1,4):
    for c in range(1,3):
        sh.cell(row=r,column=c).value='python'
wb.save(path)
print('data written sucessfully')


