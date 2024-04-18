import openpyxl


def total_Rows(xlpath,sheet_name):
    wb=openpyxl.load_workbook(xlpath)
    sh=wb[sheet_name]
    rows=sh.max_row
    return rows


def total_Columns(xlpath,sheet_name):
    wb=openpyxl.load_workbook(xlpath)
    sh=wb[sheet_name]
    Cols=sh.max_column
    return Cols

def read_Data(xlpath,sheet_name,r,c):
    wb = openpyxl.load_workbook(xlpath)
    sh = wb[sheet_name]
    data=sh.cell(row=r,column=c).value
    return data


def write_Data(xlpath,sheet_name, r,c,val):
    wb = openpyxl.load_workbook(xlpath)
    sh = wb[sheet_name]
    sh.cell(row=r,column=c).value=val
    wb.save(xlpath)




